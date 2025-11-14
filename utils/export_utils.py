"""
Utilities for exporting PAD dashboard data to Excel with comprehensive metadata
"""

import io
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def create_comprehensive_excel_report(
    df_historis: pd.DataFrame,
    df_proyeksi: pd.DataFrame,
    df_pkb_2025: pd.DataFrame,
    df_pkb_2026: pd.DataFrame,
    df_bbnkb_2025: pd.DataFrame,
    df_bbnkb_2026: pd.DataFrame,
    model_metrics: dict = None
) -> bytes:
    """
    Create comprehensive Excel report with multiple sheets

    Args:
        df_historis: Historical PAD data
        df_proyeksi: Projection data
        df_pkb_2025: PKB decomposition 2025
        df_pkb_2026: PKB decomposition 2026
        df_bbnkb_2025: BBNKB decomposition 2025
        df_bbnkb_2026: BBNKB decomposition 2026
        model_metrics: Optional dict with model performance metrics

    Returns:
        bytes: Excel file in bytes
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: Metadata & Information
        metadata_df = pd.DataFrame({
            'Field': [
                'Report Title',
                'Generated Date',
                'Generated Time',
                'Dashboard Version',
                'Data Period',
                'Projection Period',
                'Methodology',
                'Model Type',
                'Dataset Size',
                'Source',
                'Contact'
            ],
            'Value': [
                'PAD Dashboard - Comprehensive Analysis Report',
                datetime.now().strftime('%Y-%m-%d'),
                datetime.now().strftime('%H:%M:%S'),
                '1.0 (Enhanced)',
                '2018-2024 (7 years)',
                '2025-2026',
                'Linear Regression (OLS) + Macroeconomic Impact',
                'Simple Linear Regression per variable',
                f'{len(df_historis)} observations',
                'Bapenda Provinsi Jawa Timur',
                'dashboard@bapenda.jatimprov.go.id'
            ]
        })
        metadata_df.to_excel(writer, sheet_name='Metadata', index=False)

        # Sheet 2: Metodologi & Assumptions
        metodologi_df = pd.DataFrame({
            'Category': [
                'Model Assumptions',
                'Model Assumptions',
                'Model Assumptions',
                'Model Assumptions',
                'Limitations',
                'Limitations',
                'Limitations',
                'Projection Formula',
                'Projection Formula',
                'Scenario Bounds',
                'Data Sources',
                'Data Sources'
            ],
            'Item': [
                'Linear relationship',
                'Homoscedasticity',
                'Normality of residuals',
                'No multicollinearity (ideal)',
                'Small sample size (n=7)',
                'Risk of overfitting',
                'High variability',
                'Macro variables: trend regression vs Year',
                'Response: regression vs Macro variable',
                'Optimis = Pred × 0.95 | Moderat = Pred | Pesimis = Pred × 1.05',
                'Historical PAD data: Bapenda archives',
                'Macro variables: BPS, BI, and related agencies'
            ]
        })
        metodologi_df.to_excel(writer, sheet_name='Metodologi', index=False)

        # Sheet 3: Historical Data
        df_historis.to_excel(writer, sheet_name='Data Historis 2018-2024', index=False)

        # Sheet 4: Projections
        if df_proyeksi is not None and not df_proyeksi.empty:
            df_proyeksi.to_excel(writer, sheet_name='Proyeksi 2025-2026', index=False)

        # Sheet 5: PKB Decomposition 2025
        df_pkb_2025.to_excel(writer, sheet_name='PKB Dekomposisi 2025', index=False)

        # Sheet 6: PKB Decomposition 2026
        df_pkb_2026.to_excel(writer, sheet_name='PKB Dekomposisi 2026', index=False)

        # Sheet 7: BBNKB Decomposition 2025
        df_bbnkb_2025.to_excel(writer, sheet_name='BBNKB Dekomposisi 2025', index=False)

        # Sheet 8: BBNKB Decomposition 2026
        df_bbnkb_2026.to_excel(writer, sheet_name='BBNKB Dekomposisi 2026', index=False)

        # Sheet 9: Model Metrics (if provided)
        if model_metrics:
            metrics_df = pd.DataFrame([
                {'Metric': k, 'Value': v} for k, v in model_metrics.items()
            ])
            metrics_df.to_excel(writer, sheet_name='Model Performance', index=False)

        # Sheet 10: Glossary
        glossary_df = pd.DataFrame({
            'Term': [
                'PKB',
                'BBNKB',
                'PDRB',
                'Rasio Gini',
                'IPM',
                'TPT',
                'R²',
                'p-value',
                'OLS',
                'CI',
                'HKPD',
                'R-APBD'
            ],
            'Definition': [
                'Pajak Kendaraan Bermotor',
                'Bea Balik Nama Kendaraan Bermotor',
                'Produk Domestik Regional Bruto (Economic Growth)',
                'Income Inequality Index (0-1, higher = more unequal)',
                'Indeks Pembangunan Manusia (Human Development Index)',
                'Tingkat Pengangguran Terbuka (Unemployment Rate)',
                'R-Squared: Proportion of variance explained by model (0-1)',
                'Statistical significance probability (< 0.05 = significant)',
                'Ordinary Least Squares: Linear regression method',
                'Confidence Interval: Range where true value likely falls',
                'Hasil Kemitraan Pemerintah Daerah (Revenue sharing)',
                'Rencana Anggaran Pendapatan dan Belanja Daerah (Budget target)'
            ]
        })
        glossary_df.to_excel(writer, sheet_name='Glossary', index=False)

        # Apply formatting to all sheets
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]

            # Header formatting
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

    output.seek(0)
    return output.read()


def create_simple_export(df: pd.DataFrame, sheet_name: str = 'Data') -> bytes:
    """
    Create simple single-sheet Excel export

    Args:
        df: DataFrame to export
        sheet_name: Name of the sheet

    Returns:
        bytes: Excel file in bytes
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Basic formatting
        ws = writer.sheets[sheet_name]
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

    output.seek(0)
    return output.read()
